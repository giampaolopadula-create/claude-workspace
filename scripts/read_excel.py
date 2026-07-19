#!/usr/bin/env python3
"""Estrae testo da file .xlsx"""

import sys
import io
from pathlib import Path
from openpyxl import load_workbook

def read_excel(file_path):
    """Legge un file .xlsx e stampa il contenuto"""
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        wb = load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n=== Foglio: {sheet_name} ===\n")

            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    print(row_text)
    except Exception as e:
        print(f"Errore: {e}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python read_excel.py <path-to-xlsx>")
        sys.exit(1)

    read_excel(sys.argv[1])
